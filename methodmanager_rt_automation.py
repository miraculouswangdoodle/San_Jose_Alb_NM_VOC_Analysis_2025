"""
MethodManager RT Window Automation Script
==========================================

This script automates the process of updating RT Min and RT Max values in 
MethodManager's Edit Substances Table from an Excel file.

Version 4.0 Updates:
- Added logic to skip compounds where both New Rt MIN and New Rt MAX are 0
- Prints status messages for both skipped and processed compounds

Requirements:
- MethodManager software
- Python libraries: pyautogui, openpyxl, keyboard (optional)
- Excel file with RT window data (C2C6 or C6C12 format)

Author: Jaimie Ritchie
Version: 4.0
"""

import os
import sys
import time
import pyautogui
import threading
from pathlib import Path
import openpyxl
from tkinter import filedialog
import tkinter as tk

class MethodManagerAutomation:
    """Main automation class for MethodManager RT window updates"""
    
    def __init__(self):
        self.stop_requested = False
        self.compound_type = None
        self.excel_file = None
        self.compounds_data = []
        self.total_compounds = 0
        self.processed_compounds = 0
        self.skipped_compounds = 0  # NEW: Track skipped compounds
        
        # Configure pyautogui safety settings
        pyautogui.FAILSAFE = True  # Move mouse to top-left corner to stop
        pyautogui.PAUSE = 0.3      # Default pause between actions
        
        # Compound type configurations
        self.config = {
            'C2C6': {
                'rows_to_edit': 32,
                'total_rows': 33,
                'data_start_row': 7,
                'first_compound': 'ETHANE',
                'analyzer_type': 'airmoVOC C2-C6'
            },
            'C6C12': {
                'rows_to_edit': 92,
                'total_rows': 93,
                'data_start_row': 8,
                'first_compound': 'ACETALDEHYDE',
                'analyzer_type': 'airmoVOC C6-C12'
            }
        }
        
    def setup_escape_monitoring(self):
        """Set up ESC key monitoring in a separate thread"""
        def monitor_escape():
            try:
                import keyboard
                keyboard.wait('esc')
                self.stop_requested = True
                print("\n⚠️ ESC pressed - stopping automation after current operation...")
            except ImportError:
                pass
            except Exception as e:
                print(f"Error in escape monitoring: {e}")
        
        try:
            import keyboard
            escape_thread = threading.Thread(target=monitor_escape, daemon=True)
            escape_thread.start()
            print("✓ ESC key monitoring active (press ESC to stop automation)")
        except ImportError:
            print("⚠️ 'keyboard' library not found. ESC monitoring disabled.")
            print("   Install with: pip install keyboard")
    
    def select_compound_type(self):
        """Prompt user to select compound type"""
        print("\n" + "="*60)
        print("COMPOUND TYPE SELECTION")
        print("="*60)
        print("Please select the compound type:")
        print("  1. C2C6 (32 compounds: ETHANE to ...)")
        print("  2. C6C12 (92 compounds: ACETALDEHYDE to ...)")
        print()
        
        while True:
            choice = input("Enter 1 or 2: ").strip()
            if choice == '1':
                self.compound_type = 'C2C6'
                print(f"✓ Selected: C2C6")
                return True
            elif choice == '2':
                self.compound_type = 'C6C12'
                print(f"✓ Selected: C6C12")
                return True
            else:
                print("❌ Invalid choice. Please enter 1 or 2.")
    
    def select_excel_file(self):
        """Prompt user to select Excel file"""
        try:
            root = tk.Tk()
            root.withdraw()
            
            print("\n📂 Please select the Excel file with RT window data...")
            file_path = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
                initialdir=os.path.expanduser("~")
            )
            
            root.destroy()
            
            if not file_path:
                print("❌ No file selected. Exiting...")
                return False
            
            self.excel_file = file_path
            print(f"✓ Selected file: {os.path.basename(file_path)}")
            return True
            
        except Exception as e:
            print(f"Error selecting Excel file: {e}")
            return False
    
    def validate_excel_file(self):
        """Validate Excel file structure and content"""
        try:
            print("\n🔍 Validating Excel file...")
            
            # Load workbook
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws = wb.active
            
            # Check sheet name contains compound type
            sheet_name = ws.title.upper()
            print(f"📄 Active sheet: {ws.title}")
            
            if self.compound_type not in sheet_name:
                print(f"⚠️ WARNING: Sheet name doesn't contain '{self.compound_type}'")
                print(f"   Sheet name: {ws.title}")
                response = input("   Continue anyway? (y/n): ").strip().lower()
                if response != 'y':
                    wb.close()
                    return False
            
            # Find Name, RT Min, and RT Max columns in row 3
            name_col = None
            rt_min_col = None
            rt_max_col = None
            
            for col in range(1, ws.max_column + 1):
                header = ws.cell(3, col).value
                if header:
                    header_upper = str(header).upper()
                    
                    if 'NAME' in header_upper and name_col is None:
                        name_col = col
                    elif 'RT' in header_upper and 'MIN' in header_upper:
                        rt_min_col = col
                    elif 'RT' in header_upper and 'MAX' in header_upper:
                        rt_max_col = col
            
            if not name_col:
                print(f"❌ ERROR: 'Name' column not found in row 3")
                return False
            
            if not rt_min_col:
                print(f"❌ ERROR: 'New Rt MIN' column not found in row 3")
                return False
            
            if not rt_max_col:
                print(f"❌ ERROR: 'New Rt MAX' column not found in row 3")
                return False
            
            print(f"✓ Found headers: Name (col {openpyxl.utils.get_column_letter(name_col)}), RT Min (col {openpyxl.utils.get_column_letter(rt_min_col)}), RT Max (col {openpyxl.utils.get_column_letter(rt_max_col)})")
            
            # Read compound data
            config = self.config[self.compound_type]
            data_start_row = config['data_start_row']
            rows_to_read = config['rows_to_edit']
            
            self.compounds_data = []
            for i in range(rows_to_read):
                row_num = data_start_row + i
                name = ws.cell(row_num, name_col).value
                rt_min = ws.cell(row_num, rt_min_col).value
                rt_max = ws.cell(row_num, rt_max_col).value
                
                if name:  # Only add if compound name exists
                    self.compounds_data.append({
                        'name': str(name).strip(),
                        'rt_min': rt_min,
                        'rt_max': rt_max,
                        'row': row_num
                    })
            
            self.total_compounds = len(self.compounds_data)
            
            if self.total_compounds == 0:
                print(f"❌ ERROR: No compound data found starting at row {data_start_row}")
                return False
            
            # Check first compound
            first_compound = self.compounds_data[0]['name'].upper()
            expected_first = config['first_compound'].upper()
            
            if first_compound != expected_first:
                print(f"⚠️ WARNING: First compound mismatch!")
                print(f"   Expected: {expected_first}")
                print(f"   Found: {first_compound}")
                response = input("   Continue anyway? (y/n): ").strip().lower()
                if response != 'y':
                    wb.close()
                    return False
            
            print(f"✓ Found {self.total_compounds} compounds")
            wb.close()
            return True
            
        except Exception as e:
            print(f"❌ Error validating Excel file: {e}")
            return False
    
    def verify_software_setup(self):
        """Verify MethodManager is set up correctly"""
        print("\n" + "="*60)
        print("METHODMANAGER SETUP VERIFICATION")
        print("="*60)
        config = self.config[self.compound_type]
        print(f"\n📋 Please verify the following in MethodManager:")
        print(f"   1. MethodManager is open")
        print(f"   2. '{config['analyzer_type']}' analyzer is selected")
        print(f"   3. 'Edit substances table' window is open")
        print(f"   4. The table shows {config['total_rows']} rows")
        print(f"   5. First compound is: {config['first_compound']}")
        print()
        
        response = input("Is everything set up correctly? (y/n): ").strip().lower()
        return response == 'y'
    
    def _is_zero(self, value):
        """Helper function to check if a value is 0 (handles strings and numbers)"""
        if value is None or value == '':
            return False
        try:
            return float(value) == 0
        except (ValueError, TypeError):
            return False
    
    def display_dry_run(self):
        """Display preview of what will be updated"""
        print("\n" + "="*60)
        print("DRY RUN PREVIEW")
        print("="*60)
        print(f"\nThe following {self.total_compounds} compounds will be processed:")
        print()
        
        # NEW: Count all compounds that will be skipped vs processed
        will_skip = 0
        will_process = 0
        
        for compound in self.compounds_data:
            if self._is_zero(compound['rt_min']) and self._is_zero(compound['rt_max']):
                will_skip += 1
            else:
                will_process += 1
        
        # Show first 5 and last 5 compounds
        preview_compounds = []
        if self.total_compounds <= 10:
            preview_compounds = self.compounds_data
        else:
            preview_compounds = self.compounds_data[:5] + self.compounds_data[-5:]
        
        for i, compound in enumerate(preview_compounds):
            # NEW: Check if compound should be skipped
            should_skip = (self._is_zero(compound['rt_min']) and self._is_zero(compound['rt_max']))
            
            if i == 5 and self.total_compounds > 10:
                print("   ...")
            
            if should_skip:
                print(f"   {compound['name']}: SKIP (RT values are 0)")
            else:
                print(f"   {compound['name']}: RT Min={compound['rt_min']}, RT Max={compound['rt_max']}")
        
        # Check for missing values in compounds that will be processed
        missing_values = []
        for compound in self.compounds_data:
            if not (self._is_zero(compound['rt_min']) and self._is_zero(compound['rt_max'])):  # Only check non-skipped
                if compound['rt_min'] is None or compound['rt_min'] == '':
                    missing_values.append(f"   {compound['name']}: Missing RT Min value")
                if compound['rt_max'] is None or compound['rt_max'] == '':
                    missing_values.append(f"   {compound['name']}: Missing RT Max value")
        
        if missing_values:
            print("\n⚠️ WARNING: Missing RT values detected:")
            for msg in missing_values:
                print(msg)
        
        # NEW: Display skip/process summary
        print("\n" + "="*60)
        print(f"Summary:")
        print(f"  Total compounds: {self.total_compounds}")
        print(f"  Will be PROCESSED: {will_process}")
        print(f"  Will be SKIPPED: {will_skip}")
        print("="*60)
        
        response = input("\nProceed with automation? (y/n): ").strip().lower()
        return response == 'y'
    
    def prepare_for_data_entry(self):
        """Guide user to prepare for data entry"""
        print("\n" + "="*60)
        print("DATA ENTRY PREPARATION")
        print("="*60)
        print("\n📋 Please do the following:")
        print("   1. Click on the MethodManager 'Edit substances table' window")
        print("   2. Click on the FIRST COMPOUND NAME cell in the table")
        config = self.config[self.compound_type]
        print(f"      (Should be: {config['first_compound']})")
        print("   3. Make sure the cell is SELECTED (not in edit mode)")
        print("   4. Do NOT move mouse or keyboard after pressing Enter")
        print()
        print("⚠️ IMPORTANT: The automation will start immediately after Enter!")
        print("After automation starts, click on the edit substances table window to allow the automation to operate")
        print()
        input("Press Enter when the first compound name cell is selected...")
        
        print("\n🚀 Starting automation in 3 seconds...")
        time.sleep(1)
        print("   2...")
        time.sleep(1)
        print("   1...")
        time.sleep(1)
        print("   GO!")
    
    def update_rt_values(self):
        """Main data entry automation loop"""
        try:
            print("\n" + "="*60)
            print("UPDATING RT VALUES")
            print("="*60)
            
            for i, compound in enumerate(self.compounds_data):
                if self.stop_requested:
                    print("\n⚠️ Automation stopped by user")
                    return False
                
                # NEW: Check if compound should be skipped
                should_skip = (self._is_zero(compound['rt_min']) and self._is_zero(compound['rt_max']))
                
                if should_skip:
                    # NEW: Print skip message and move to next compound
                    print(f"\nSkipping {i+1}/{self.total_compounds}: {compound['name']} - RT values are 0")
                    self.skipped_compounds += 1
                    
                    # Move to next compound's Name cell (down arrow)
                    pyautogui.press('down')
                    time.sleep(0.2)
                    
                    continue  # Skip to next compound
                
                # Process compound normally
                print(f"\nProcessing {i+1}/{self.total_compounds}: {compound['name']}")
                
                # Current position: Name cell
                # Move to RT Min cell
                pyautogui.press('right')
                time.sleep(0.2)
                
                # Enter edit mode
                pyautogui.press('enter')
                time.sleep(0.2)
                
                # Clear existing value and type new RT Min
                pyautogui.hotkey('ctrl', 'a')
                time.sleep(0.1)
                if compound['rt_min'] is not None:
                    pyautogui.typewrite(str(compound['rt_min']), interval=0.05)
                time.sleep(0.2)
                
                # Exit edit mode
                pyautogui.press('enter')
                time.sleep(0.2)
                
                # Move to RT Max cell
                pyautogui.press('right')
                time.sleep(0.2)
                
                # Enter edit mode
                pyautogui.press('enter')
                time.sleep(0.2)
                
                # Clear existing value and type new RT Max
                pyautogui.hotkey('ctrl', 'a')
                time.sleep(0.1)
                if compound['rt_max'] is not None:
                    pyautogui.typewrite(str(compound['rt_max']), interval=0.05)
                time.sleep(0.2)
                
                # Exit edit mode
                pyautogui.press('enter')
                time.sleep(0.2)
                
                # Move to next compound's Name cell
                # Down arrow (moves to next row, RT Max column)
                pyautogui.press('down')
                time.sleep(0.2)
                
                # Left arrow twice (RT Max -> RT Min -> Name)
                pyautogui.press('left')
                time.sleep(0.2)
                pyautogui.press('left')
                time.sleep(0.2)
                
                self.processed_compounds += 1
                print(f"  ✓ Updated RT Min: {compound['rt_min']}, RT Max: {compound['rt_max']}")
            
            print("\n✓ All compounds processed successfully!")
            return True
            
        except Exception as e:
            print(f"\n❌ Error during data entry: {e}")
            print(f"   Last processed: {self.processed_compounds}/{self.total_compounds}")
            return False
    
    def print_final_report(self):
        """Print final automation report"""
        print("\n" + "="*60)
        print("AUTOMATION COMPLETE")
        print("="*60)
        print(f"\nCompound Type: {self.compound_type}")
        print(f"Total Compounds: {self.total_compounds}")
        print(f"Successfully Processed: {self.processed_compounds}")
        print(f"Skipped (RT=0): {self.skipped_compounds}")  # NEW: Show skipped count
        
        if self.processed_compounds + self.skipped_compounds < self.total_compounds:
            incomplete = self.total_compounds - self.processed_compounds - self.skipped_compounds
            print(f"⚠️ Incomplete: {incomplete} compounds not processed")
        else:
            print("✓ All compounds handled successfully!")
        
        print("\n📋 Next Steps:")
        print("   1. Verify the RT values in MethodManager")
        print("   2. Save the substance table if not auto-saved")
        print("   3. Close or continue working in MethodManager")
        print("="*60)
    
    def run_automation(self):
        """Main automation workflow"""
        try:
            print("="*60)
            print("🚀 METHODMANAGER RT WINDOW AUTOMATION")
            print("="*60)
            print("Press ESC at any time to stop the automation safely")
            print()
            
            # Phase 1: Setup escape monitoring
            self.setup_escape_monitoring()
            
            # Phase 2: Select compound type
            if not self.select_compound_type():
                return
            
            if self.stop_requested:
                return
            
            # Phase 3: Select Excel file
            if not self.select_excel_file():
                return
            
            if self.stop_requested:
                return
            
            # Phase 4: Validate Excel file
            if not self.validate_excel_file():
                return
            
            if self.stop_requested:
                return
            
            # Phase 5: Verify MethodManager setup
            if not self.verify_software_setup():
                return
            
            if self.stop_requested:
                return
            
            # Phase 6: Dry run preview
            if not self.display_dry_run():
                print("\n❌ Automation cancelled by user")
                return
            
            if self.stop_requested:
                return
            
            # Phase 7: Prepare for data entry
            self.prepare_for_data_entry()
            
            if self.stop_requested:
                return
            
            # Phase 8: Execute data entry
            self.update_rt_values()
            
            # Phase 9: Final report
            self.print_final_report()
            
        except Exception as e:
            print(f"\n❌ Critical error in automation: {e}")
            print("❌ Please check the error message above and try again.")

def main():
    """Main function to run the automation"""
    try:
        print("Starting MethodManager RT Window Automation...")
        automation = MethodManagerAutomation()
        automation.run_automation()
    except KeyboardInterrupt:
        print("\n\n⚠️ Automation interrupted by user (Ctrl+C)")
    except ImportError as e:
        print(f"\n❌ Missing required library: {e}")
        print("Please install required libraries with: pip install pyautogui openpyxl keyboard")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        print("Please check that all required libraries are installed.")
    finally:
        print("\n" + "="*60)
        print("🏁 SCRIPT FINISHED")
        print("="*60)
        input("Press Enter to exit...")

if __name__ == "__main__":
    main()
